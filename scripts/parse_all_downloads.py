import os
import zipfile
import xml.etree.ElementTree as ET
import openpyxl

downloads = os.path.expanduser('~/Downloads')

target_files = [
    '(10-26-1) 1 Guía-Trabajo Final.docx',
    '(10-26-1) 1 Guía-Trabajo Final.pdf',
    '(10-26-1) 2 Instrumento de Evaluación.xlsx',
    '(10-26-2) 3 Modelo de Paper.pdf',
    '10-26-1) Medición AG.docx',
    '251011 Informe de Derechos Autor.docx',
    'AG-C05_ Gestión de Proyectos_Vera_de_la_Cruz_Nilton_Alonso.pdf',
    'Ficha de Evaluación Soft. 2025-02.docx'
]

out_log = os.path.join(os.path.dirname(__file__), 'downloads_parsed_summary.txt')

with open(out_log, 'w', encoding='utf-8') as out:
    for fname in target_files:
        fpath = os.path.join(downloads, fname)
        out.write(f"\n=========================================\nFILE: {fname}\n=========================================\n")
        if not os.path.exists(fpath):
            out.write("FILE NOT FOUND!\n")
            continue
        
        if fname.endswith('.docx'):
            try:
                with zipfile.ZipFile(fpath) as z:
                    xml_content = z.read('word/document.xml')
                    tree = ET.fromstring(xml_content)
                    texts = [node.text for node in tree.iter() if node.text]
                    out.write(' '.join(texts) + '\n')
            except Exception as e:
                out.write(f"Error reading docx: {e}\n")
                
        elif fname.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                for sheetname in wb.sheetnames:
                    out.write(f"\n--- SHEET: {sheetname} ---\n")
                    ws = wb[sheetname]
                    for row in ws.iter_rows(values_only=True):
                        row_vals = [str(cell) for cell in row if cell is not None]
                        if row_vals:
                            out.write(' | '.join(row_vals) + '\n')
            except Exception as e:
                out.write(f"Error reading xlsx: {e}\n")
                
        elif fname.endswith('.pdf'):
            try:
                with open(fpath, 'rb') as pf:
                    content = pf.read()
                    # extract plain ASCII/UTF-8 strings from PDF
                    import re
                    # find text objects or raw strings
                    strings = re.findall(rb'\((.*?)\)', content)
                    decoded = []
                    for s in strings:
                        try:
                            decoded.append(s.decode('utf-8', errors='ignore'))
                        except:
                            pass
                    out.write(' '.join(decoded[:5000]) + '\n')
            except Exception as e:
                out.write(f"Error reading pdf: {e}\n")

print(f"Extraction completed! Saved to {out_log}")
