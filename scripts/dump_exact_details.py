import os
import sys
import zipfile
import xml.etree.ElementTree as ET
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
downloads = os.path.expanduser('~/Downloads')

def dump_docx(fname):
    fpath = os.path.join(downloads, fname)
    print(f"\n=========================================\nFULL DOCX: {fname}\n=========================================")
    with zipfile.ZipFile(fpath) as z:
        xml_content = z.read('word/document.xml')
        tree = ET.fromstring(xml_content)
        # get all paragraph texts
        paragraphs = []
        for p in tree.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
            p_text = ''.join([node.text for node in p.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t') if node.text])
            if p_text.strip():
                paragraphs.append(p_text.strip())
        print('\n'.join(paragraphs))

def dump_xlsx(fname):
    fpath = os.path.join(downloads, fname)
    print(f"\n=========================================\nFULL XLSX: {fname}\n=========================================")
    wb = openpyxl.load_workbook(fpath, data_only=True)
    for sheet in wb.sheetnames:
        print(f"\n--- SHEET: {sheet} ---")
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) for c in row if c is not None]
            if vals:
                print(' | '.join(vals))

print("DUMPING EXACT DOCX & XLSX CONTENTS:")
dump_docx('(10-26-1) 1 Guía-Trabajo Final.docx')
dump_docx('10-26-1) Medición AG.docx')
dump_docx('251011 Informe de Derechos Autor.docx')
dump_docx('Ficha de Evaluación Soft. 2025-02.docx')
dump_xlsx('(10-26-1) 2 Instrumento de Evaluación.xlsx')
